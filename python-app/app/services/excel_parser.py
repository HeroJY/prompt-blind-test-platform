# -*- coding: utf-8 -*-

from io import BytesIO

from openpyxl import load_workbook


def validate_excel_filename(filename):
    # type: (str) -> str
    if not filename:
        return "file name is required"

    lower_name = filename.lower()
    if not lower_name.endswith(".xlsx"):
        return "only .xlsx upload is supported now"

    return ""


def _read_rows(file_bytes):
    # type: (bytes) -> list
    workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("excel is empty")
    return rows


def _normalize_headers(header_row):
    # type: (tuple) -> list
    headers = []
    for value in header_row:
        if value is None:
            headers.append("")
        else:
            headers.append(str(value).strip().lower())
    return headers


def _row_to_dict(headers, row):
    # type: (list, tuple) -> dict
    data = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        value = row[index] if index < len(row) else None
        if value is None:
            data[header] = ""
        else:
            data[header] = str(value).strip()
    return data


def _non_empty_dicts(rows):
    # type: (list) -> list
    if len(rows) < 2:
        return []

    headers = _normalize_headers(rows[0])
    result = []
    for row in rows[1:]:
        item = _row_to_dict(headers, row)
        if any(item.values()):
            result.append(item)
    return result


def parse_prompt_excel(file_bytes):
    # type: (bytes) -> dict
    rows = _read_rows(file_bytes)
    items = _non_empty_dicts(rows)

    if not items:
        raise ValueError("prompt excel has no data rows")

    first_item = items[0]
    required_fields = ["task_name", "task_description", "prompt_a", "prompt_b"]
    for field in required_fields:
        if not first_item.get(field):
            raise ValueError("prompt excel missing field: {0}".format(field))

    return {
        "task_name": first_item.get("task_name", ""),
        "task_description": first_item.get("task_description", ""),
        "prompt_a_text": first_item.get("prompt_a", ""),
        "prompt_b_text": first_item.get("prompt_b", ""),
    }


def parse_item_excel(file_bytes):
    # type: (bytes) -> dict
    rows = _read_rows(file_bytes)
    items = _non_empty_dicts(rows)

    if not items:
        raise ValueError("item excel has no data rows")

    parsed_items = []
    for index, item in enumerate(items):
        code = item.get("code", "")
        source_text = item.get("source_text", "")
        if not code:
            raise ValueError("row {0} missing code".format(index + 2))
        if not source_text:
            raise ValueError("row {0} missing source_text".format(index + 2))

        parsed_items.append(
            {
                "code": code,
                "sort_order": int(item.get("sort_order", index + 1)),
                "source_type": item.get("source_type", "text") or "text",
                "source_text": source_text,
            }
        )

    return {
        "item_count": len(parsed_items),
        "items": parsed_items,
    }


def parse_test_data_excel(file_bytes):
    # type: (bytes) -> dict
    rows = _read_rows(file_bytes)
    items = _non_empty_dicts(rows)

    if not items:
        raise ValueError("test data excel has no data rows")

    preview_lines = []
    for index, item in enumerate(items):
        data_key = item.get("data_key", "")
        data_value = item.get("data_value", "")
        if not data_key:
            raise ValueError("row {0} missing data_key".format(index + 2))
        preview_lines.append("{0}={1}".format(data_key, data_value))

    return {
        "row_count": len(items),
        "preview_text": "\n".join(preview_lines),
    }
